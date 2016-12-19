from openpyxl import *
from data import *
import os


cols = ['No.', 'ID', 'Gene', 'Transcript RefSeq ID', 'Variation Location',
        'Gencode Transcript', 'Exons', 'Variation Type', 'Molecular Conseq', 'Translocation Name',
        'Nucleotide Chagne', 'Protein change', 'Conditions', 'FATHMM score', 'FDA-approved indication(s)',
        'Agent', 'Ref.', 'Criteria/Study ID', 'Note/Study ref.']


def load_xlsx(xlsx_path):
    wb = None
    if os.path.isfile(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        sheet = wb.create_sheet('data')
        for i, v in enumerate(cols):
            sheet.cell(row=1, column=i + 1).value = v
    return wb


def write_misses(output_path):
    file = open(output_path + '_misses.txt', 'a')
    for miss in misses:
        file.write(miss + '\n')


def load_references(reference_path):
    if os.path.isfile(reference_path):
        with open(reference_path) as f:
            for line in f.readlines():
                if line.strip() == '':
                    continue
                index, title = line.split('\t', 1)
                references_dict[title] = int(index)


def write_references(reference_path):
    file = open(reference_path, 'a')
    references = [str] * len(references_dict)
    for key, value in references_dict.items():
        references[value] = key
    for index, reference in enumerate(references):
        file.write('{}\t{}\n'.format(index, reference))
    file.close()


def write(wb, xlsx_path):
    wb.save(xlsx_path)


def get_gene_list(gene_path):
    with open(gene_path) as f:
        return [(i + 1, x.strip()) for i, x in enumerate(f.readlines())]


def get_genes(input_path):
    genes = []
    wb = load_workbook(input_path)
    genes_sheet = wb.get_sheet_by_name('Ref.')
    for r in range(5, genes_sheet.max_row):
        index = genes_sheet.cell(row=r, column=1).value
        gen = genes_sheet.cell(row=r, column=2).value
        genes.append((index, gen))
    return genes


def output(wb, data):
    sheet = wb.get_sheet_by_name('data')
    row = sheet.max_row + 1

    def standard():
        cell(2, data.gen_id)
        cell(3, data.gen)

    def cell(col, val):
        nonlocal sheet
        nonlocal row
        sheet.cell(row=row, column=col).value = val

    cell(1, data.index)

    for tissue in data.tissues:
        standard()
        cell(8, tissue.variation_type)
        cell(9, tissue.conseq)
        hnscc = 'HNSCC ({:.1%}, {}/{})'.format(float(tissue.value)/tissue.total, tissue.value, tissue.total)
        cell(13, hnscc)
        row += 1
    for fusion in data.fusions:
        standard()
        cell(8, 'Fusion')
        cell(9, '-')
        cell(10, fusion.mutation)
        cell(17, fusion.link)
        cell(18, 'count=' + str(fusion.count))
        row += 1
    for elem in combined_list(data):
        standard()
        var_loc = '{} {} / {}'.format(data.gen, elem.aa_mut, elem.cds_mut)
        cell(5, var_loc)
        cell(6, elem.transcript)
        cell(9, elem.conseq)
        cell(11, elem.cds_mut)
        cell(12, elem.aa_mut.split('.', 1)[1])
        cell(13, '  '.join(map(str, elem.references)))
        cell(17, elem.link)
        cell(18, 'count=' + str(elem.count))
        if elem.type == 0:
            cell(8, elem.variation_type)
        elif elem.type == 1:
            cell(8, 'SNV')
            cell(14, elem.score)
        row += 1


def combined_list(data):
    l = []
    for key, snv in data.snvs.items():
        l.append(snv)
    for key, frameshift in data.frameshifts.items():
        l.append(frameshift)
    l.sort(key=lambda x: x.order_mut)
    return l
